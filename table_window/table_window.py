from itertools import groupby

import matplotlib.pyplot as plt
import numpy as np
import openpyxl as opl
from PyQt5 import QtWidgets
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

import definitions
import table_model
from .table_window_design import Ui_Form


class TableWindow(QtWidgets.QWidget, Ui_Form):
    def __init__(self, dataframe, description):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле main_window_design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна

        self.label_spends.setText(self.label_spends.text()
                                  + '\n'.join([f'{key} - {value}' for key, value in definitions.SPENDS_NAMES.items()]))

        self._description = description
        self.label_description.setText(self._description)

        dataframe['Коэффициент'] = np.ones(dataframe.shape[0])
        dataframe.loc['ИТОГО'] = np.zeros(dataframe.shape[1])

        self._columns_names = dataframe.columns.tolist()
        self._rows_names = dataframe.index.tolist()

        self._model = table_model.TableModel(dataframe)
        self.tableView.setModel(self._model)
        self._setup_stretch_and_color()

        self.btn_save_table.clicked.connect(lambda: self._handle_btn_save_table(
            "Сохранение таблицы"))
        self.btn_build_graphs.clicked.connect(self._handle_btn_build_graphs)

    def _setup_stretch_and_color(self):
        h_header = self.tableView.horizontalHeader()
        for i in range(len(self._columns_names)):
            h_header.setSectionResizeMode(i, QtWidgets.QHeaderView.Stretch)
        v_header = self.tableView.verticalHeader()
        for i in range(len(self._rows_names)):
            v_header.setSectionResizeMode(i, QtWidgets.QHeaderView.Stretch)

        self.tableView.setStyleSheet("QHeaderView::section { background-color:rgb"
                                     + str(definitions.HEADER_COLOR_RGB) + " }")

    def _handle_btn_build_graphs(self):
        selected_indexes = self.tableView.selectedIndexes()

        selected_indexes = [(el.row(), el.column()) for el in selected_indexes
                            if el.column() != (len(self._columns_names) - 1)]

        if len(selected_indexes) == 0:
            QtWidgets.QMessageBox.critical(self, "Ошибка", "Выберите ячейки")
            return

        def first(x):
            return x[0]

        def second(x):
            return x[1]

        coefficients = self._model.get_coefficients()

        if self.comboBox_row_col.currentIndex() == 0:  # by rows
            key_sort_groupby = first

            graph_titles = [f"{definitions.SPENDS_NAMES[self._rows_names[i]]} (x{coefficients[i]})"
                            for i in range(len(self._rows_names[:-1]))]  # без ИТОГО
            graph_titles.append(self._rows_names[-1])

            key_indexes = second
            x_axis_labels = self._columns_names

        else:  # by columns
            key_sort_groupby = second
            graph_titles = self._columns_names
            key_indexes = first

            x_axis_labels = [f"{self._rows_names[i]} (x{coefficients[i]})"
                             for i in range(len(self._rows_names[:-1]))]
            x_axis_labels.append(self._rows_names[-1])

        selected_indexes.sort(key=key_sort_groupby)
        for graph_index, coordinates_tuples in groupby(selected_indexes, key=key_sort_groupby):
            # нарисовать очередной график
            coordinates_tuples = list(coordinates_tuples)

            xs = [x_axis_labels[i] for i in map(key_indexes, coordinates_tuples)]
            ys = [self._model.get_cell_value(i, j) for i, j in coordinates_tuples]

            self._show_graph(xs, ys, graph_titles[graph_index])

        plt.show()

    def _show_graph(self, xs, ys, title):
        fig = plt.figure()
        fig.canvas.set_window_title(title)
        plt.bar(xs, ys)
        plt.grid()
        plt.ylabel("Рублей (тыс.)")
        dy = max(ys) * 0.01  # shift of the label over each column
        for index in range(len(ys)):
            plt.text(index, ys[index] + dy, definitions.format_number(ys[index]), horizontalalignment='center')

        plt.title(title + '\n' + self._description.replace(";", ";\n"))

    def _handle_btn_save_table(self, prompt):
        # Данный метод сохраняет выведенную на экран таблицу (в table_window) по нажатию на кнопку сохранить
        filename = QtWidgets.QFileDialog.getSaveFileName(self, prompt, "", "Excel-файлы (*.xlsx)")
        if filename[0] == "":
            return
        # print(f'"{filename[0]}"')
        try:
            # save dataframe as is
            self._model.get_output_dataframe().to_excel(filename[0])

            # Открытие таблицы
            wb = opl.load_workbook(filename[0])
            # Выбираем лист (он у нас один [0])
            sheet = wb.worksheets[0]

            # Подгонка ширины столбцов
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0),
                                                        self._get_cell_length(cell.value)))
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value * (1 if col == 'A' else 1.2)

            # Координаты ячеек для объединения (в них будет записано описание)
            coordinates = (f"A{sheet.max_row + 1}", f"{get_column_letter(sheet.max_column)}{sheet.max_row + 1}")
            sheet[coordinates[0]] = self._description  # вставляем описание таблицы в excel
            sheet.merge_cells(f'{coordinates[0]}:{coordinates[1]}')

            for cell in sheet['A']:
                cell.alignment = Alignment(horizontal='left')

            # Выравнивание по центру
            sheet[coordinates[0]].alignment = Alignment(horizontal='center')

            # Перезапись файла
            wb.save(filename[0])

        except Exception as e:
            print("Ошибка сохранения таблицы!")
            print(e)

    @staticmethod
    def _get_cell_length(value):
        value = str(value)
        try:
            float_value = float(value)
            str_value = str(round(float_value, definitions.DIGITS_AFTER_DECIMAL))
        except ValueError:
            str_value = value
        return len(str_value)
